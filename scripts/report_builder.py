"""
Generic report renderer used by generate_exports.py.

A "report" is a dict:
{
    "title": str,
    "subtitle": str,
    "snapshot_date": "YYYY-MM-DD",
    "blocks": [
        {"type": "table", "heading": str, "headers": [...], "rows": [[...], ...],
         "col_widths": [optional, for PDF],
         "note": optional str shown under the table},
        {"type": "text", "heading": str, "body": str},
    ]
}
"""
from pathlib import Path
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors as rl_colors_mod
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from branding import TM_BLUE, TM_MAROON, TM_GRAY, TM_LIGHT_GRAY


def hexc(h):
    return rl_colors_mod.HexColor(f"#{h}")


def build_pdf(report, out_path):
    doc = SimpleDocTemplate(
        str(out_path), pagesize=letter,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TMTitle", parent=styles["Title"], textColor=hexc(TM_BLUE),
        fontSize=18, spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "TMSubtitle", parent=styles["Normal"], textColor=hexc(TM_MAROON),
        fontSize=12, spaceAfter=2,
    )
    meta_style = ParagraphStyle(
        "TMMeta", parent=styles["Normal"], textColor=rl_colors_mod.HexColor("#555555"),
        fontSize=9, spaceAfter=14,
    )
    heading_style = ParagraphStyle(
        "TMHeading", parent=styles["Heading2"], textColor=hexc(TM_BLUE),
        fontSize=13, spaceBefore=14, spaceAfter=6,
    )
    note_style = ParagraphStyle(
        "TMNote", parent=styles["Normal"], fontSize=8.5,
        textColor=rl_colors_mod.HexColor("#555555"), spaceAfter=10, spaceBefore=4,
    )
    body_style = styles["Normal"]

    story = [
        Paragraph(report["title"], title_style),
        Paragraph(report.get("subtitle", ""), subtitle_style),
        Paragraph(
            f"District 82 Toastmasters &bull; Snapshot as of {report['snapshot_date']} "
            f"&bull; Pathways Chair Dashboard",
            meta_style,
        ),
    ]

    for block in report["blocks"]:
        if block["type"] == "table":
            story.append(Paragraph(block["heading"], heading_style))
            data = [block["headers"]] + block["rows"]
            col_widths = block.get("col_widths")
            t = Table(data, colWidths=col_widths, repeatRows=1)
            style = [
                ("BACKGROUND", (0, 0), (-1, 0), hexc(TM_BLUE)),
                ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors_mod.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.4, rl_colors_mod.HexColor("#CCCCCC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [rl_colors_mod.white, rl_colors_mod.HexColor(f"#{TM_LIGHT_GRAY}")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
            t.setStyle(TableStyle(style))
            story.append(t)
            if block.get("note"):
                story.append(Paragraph(block["note"], note_style))
            story.append(Spacer(1, 6))
        elif block["type"] == "text":
            story.append(Paragraph(block["heading"], heading_style))
            story.append(Paragraph(block["body"], body_style))
            story.append(Spacer(1, 6))

    doc.build(story)


def build_xlsx(report, out_path):
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor=TM_BLUE)
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    title_font = Font(color=TM_BLUE, bold=True, name="Arial", size=14)
    subtitle_font = Font(color=TM_MAROON, bold=True, name="Arial", size=11)
    meta_font = Font(color="555555", name="Arial", size=9, italic=True)
    body_font = Font(name="Arial", size=10)
    stripe_fill = PatternFill("solid", fgColor=TM_LIGHT_GRAY)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    used_names = set()

    for block in report["blocks"]:
        if block["type"] != "table":
            continue
        base = block["heading"][:28] or "Sheet"
        name = base
        i = 2
        while name in used_names:
            name = f"{base[:26]}_{i}"
            i += 1
        used_names.add(name)
        ws = wb.create_sheet(title=name)

        ws["A1"] = report["title"]
        ws["A1"].font = title_font
        ws["A2"] = report.get("subtitle", "")
        ws["A2"].font = subtitle_font
        ws["A3"] = f"Snapshot as of {report['snapshot_date']}  |  District 82 Toastmasters Pathways Dashboard"
        ws["A3"].font = meta_font
        ws["A4"] = block["heading"]
        ws["A4"].font = subtitle_font

        start_row = 6
        headers = block["headers"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for r, row in enumerate(block["rows"], start=start_row + 1):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = body_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
                if (r - start_row) % 2 == 0:
                    cell.fill = stripe_fill

        for c, h in enumerate(headers, start=1):
            maxlen = max([len(str(h))] + [len(str(row[c - 1])) for row in block["rows"]]) if block["rows"] else len(str(h))
            ws.column_dimensions[get_column_letter(c)].width = min(max(maxlen + 3, 10), 45)

        if block.get("note"):
            note_row = start_row + len(block["rows"]) + 2
            ws.cell(row=note_row, column=1, value=block["note"]).font = meta_font

    if not wb.sheetnames:
        wb.create_sheet("Report")
    wb.save(str(out_path))


def render_report(report, out_dir: Path, filename_base: str):
    out_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = out_dir / f"{filename_base}.pdf"
    xlsx_path = out_dir / f"{filename_base}.xlsx"
    build_pdf(report, pdf_path)
    build_xlsx(report, xlsx_path)
    return pdf_path, xlsx_path
